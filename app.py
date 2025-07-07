from flask import Flask, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import re

app = Flask(__name__)
FILENAME = "client_hunting.xlsx"

# Ensure Excel exists
if not os.path.exists(FILENAME):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Phone", "Email", "Profession", "Date"])
    wb.save(FILENAME)

def extract_leads(text):
    leads = []
    entries = text.strip().split("\n\n")  # Separate each lead by blank lines

    for entry in entries:
        name = re.search(r"(Name|Client Name):\s*(.+)", entry, re.IGNORECASE)
        phone = re.search(r"(Phone|Number):\s*(.+)", entry, re.IGNORECASE)
        email = re.search(r"(Email):\s*(.+)", entry, re.IGNORECASE)
        profession = re.search(r"(Profession):\s*(.+)", entry, re.IGNORECASE)

        leads.append([
            name.group(2).strip() if name else "",
            phone.group(2).strip() if phone else "",
            email.group(2).strip() if email else "",
            profession.group(2).strip() if profession else "",
            datetime.now().strftime("%Y-%m-%d")
        ])
    return leads

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        raw_text = request.form['lead_text']
        leads = extract_leads(raw_text)

        wb = load_workbook(FILENAME)
        ws = wb.active
        for lead in leads:
            ws.append(lead)
        wb.save(FILENAME)

        return render_template('index.html', message="✅ Leads saved successfully!", file_ready=True)

    return render_template('index.html', message="", file_ready=False)

@app.route('/download')
def download():
    return send_file(FILENAME, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
